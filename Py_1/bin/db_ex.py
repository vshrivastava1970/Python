# List Comprehension
L1 = [i for i in range(10)]  # each element will be added into list

L2 = [j for j in L1 if j % 2 == 0]

# read the file
F = open(r'D:/Training/log/log_sample.txt')

data = F.readlines()
F.close()

ip = [line.split()[0] for line in data if line[0].isdigit()]
dt = [line.split()[3].split(':')[0].strip('[') for line in data if line[0].isdigit()]
ws = [line.split()[10].split('/')[2] for line in data if line[0].isdigit()]
pics = [line.split()[6].split('/')[-1] if line.split()[6].startswith('/pics') else 'No Image' for line in data if
        line[0].isdigit()]
print(ip)
print(dt)
print(ws)
print(pics)

# Python to connect DB
import sqlite3

con = sqlite3.connect('mydb')
# similarly, you can connect to other database like:
# xyz.connect(user='', password='',host='', port= , dadabase='mydb_xyz')
# example for oracle, 'cx_Oracle' module will be available

cur = con.cursor()
cur.execute(
    'create table if not exists log_sample(ip VARCHAR(100), dt VARCHAR(100), pics VARCHAR(100), ws VARCHAR(100))')

import openpyxl

wb1 = openpyxl.Workbook()
wb2 = openpyxl.Workbook()
sheets = wb1.sheetnames
ws1 = wb1['Sheet']  # return existing sheet
ws2 = wb2.create_sheet('Report')  # create new sheet

ws1['A1'] = 100
ws2['A2'] = 100

L = [10, 20]
ws1.append(L)
ws2.append(L)

ws1.cell(row=3, column=3).value = 200
ws2.cell(row=3, column=3).value = 200

cur.execute("delete from log_sample")
for i, j, k, l in zip(ip, dt, pics, ws):
    q = "insert into log_sample values('{}','{}','{}','{}')".format(i, j, k, l)
    cur.execute(q)
con.commit()

cur.execute('select * from log_sample')
result = cur.fetchall()
print('Result=', result)

for i in result:
    ws1.append(i)
    ws2.append(i)

wb1.save('report5.xlsx')
wb2.save('report6.xlsx')

wb3 = openpyxl.load_workbook('report5.xlsx')
sheets = wb3.sheetnames
ws3 = wb3['Sheet']

tr = ws3.max_row
tc = ws3.max_column

for r in range(3, tr + 1):
    print(ws3.cell(row=r, column=1).value)

for r in range(3, tr + 1):
    for c in range(1, 4):
        print(ws3.cell(row=r, column=c).value)

for i in ws3['A']:
    print('Col-A=', i.value)

for i in ws3['3']:
    print('Row-3=', i.value)

for i in ws3['B:D']:
    #print('B to D=', i)
    for x in i:
        print(x.value, end=' ')
    print()


for i in ws3['3:5']:
    #print('B to D=', i)
    for x in i:
        print('3-5', x.value, end=' ')
    print()
# wb3 = openpyxl.load_workbook('report5.xlsx')
########################################################

from openpyxl.chart import Reference, AreaChart, BarChart

wb4 = openpyxl.Workbook()
ws41 = wb4.create_sheet('Data')
ws42 = wb4.create_sheet('Area')
ws43 = wb4.create_sheet('Bar')

for i in range(1, 10):
    for j in range(1, 5):
        ws41.cell(row=i, column=j).value =i*j
wb4.save('chart.xlsx')
wb4.close()

wb5 = openpyxl.load_workbook('chart.xlsx')
ws51 = wb5['Data']
ws52 = wb5['Area']
ws53 = wb5['Bar']

achart = AreaChart()

category = Reference(ws51, min_row=1, min_col=1, max_row=10)
data = Reference(ws51, min_row=1, min_col=2, max_row=10, max_col=4)
achart.set_categories(category)
achart.add_data(data)

ws52.add_chart(achart, 'A1')

wb5.save('chart.xlsx')

bchart = BarChart()
bchart.set_categories(category)
bchart.add_data(data)

ws53.add_chart(bchart, 'A1')

wb5.save('chart.xlsx')
